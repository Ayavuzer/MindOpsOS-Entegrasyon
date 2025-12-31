"""Excel report service for sync operations."""

from io import BytesIO
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import asyncpg


class SyncReportService:
    """Service for generating Excel reports of sync operations."""
    
    def __init__(self, pool: asyncpg.Pool):
        self.pool = pool
    
    async def generate_excel_report(self, sync_id: str, tenant_id: int) -> Optional[bytes]:
        """
        Generate an Excel report for a sync operation.
        
        Args:
            sync_id: Sync run ID
            tenant_id: Tenant ID for verification
            
        Returns:
            Excel file as bytes, or None if sync not found
        """
        # Get sync run data
        async with self.pool.acquire() as conn:
            run = await conn.fetchrow(
                """
                SELECT * FROM sync_runs 
                WHERE sync_id = $1 AND tenant_id = $2
                """,
                sync_id, tenant_id,
            )
            
            if not run:
                return None
            
            # Get sync items with email details
            items = await conn.fetch(
                """
                SELECT 
                    si.id,
                    si.email_id,
                    si.item_type,
                    si.status,
                    si.sedna_rec_id,
                    si.error_message,
                    si.processed_at,
                    e.subject,
                    e.sender,
                    e.received_at,
                    e.voucher_no
                FROM sync_items si
                JOIN emails e ON si.email_id = e.id
                WHERE si.sync_run_id = $1
                ORDER BY si.id
                """,
                run["id"],
            )
        
        # Create workbook
        wb = Workbook()
        
        # =======================================================================
        # Summary Sheet
        # =======================================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # Title
        ws_summary["A1"] = "Sedna Sync Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary.merge_cells("A1:D1")
        
        # Summary data
        summary_data = [
            ("Sync ID:", sync_id),
            ("Status:", run["status"].capitalize()),
            ("Started At:", run["started_at"].strftime("%Y-%m-%d %H:%M:%S") if run["started_at"] else "-"),
            ("Completed At:", run["completed_at"].strftime("%Y-%m-%d %H:%M:%S") if run["completed_at"] else "-"),
            ("", ""),
            ("Total Items:", run["total_items"]),
            ("Successful:", run["successful_count"]),
            ("Failed:", run["failed_count"]),
            ("Success Rate:", f"{(run['successful_count'] / run['total_items'] * 100):.1f}%" if run["total_items"] > 0 else "0%"),
        ]
        
        for i, (label, value) in enumerate(summary_data, start=3):
            ws_summary[f"A{i}"] = label
            ws_summary[f"A{i}"].font = Font(bold=True)
            ws_summary[f"B{i}"] = value
        
        # Set column widths
        ws_summary.column_dimensions["A"].width = 15
        ws_summary.column_dimensions["B"].width = 25
        
        # =======================================================================
        # All Items Sheet
        # =======================================================================
        ws_all = wb.create_sheet("All Items")
        
        # Headers
        headers = ["#", "Email ID", "Subject", "Sender", "Type", "Status", "Sedna ID", "Error", "Processed At"]
        for col, header in enumerate(headers, start=1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Data
        for row, item in enumerate(items, start=2):
            ws_all.cell(row=row, column=1, value=row - 1).border = thin_border
            ws_all.cell(row=row, column=2, value=item["email_id"]).border = thin_border
            ws_all.cell(row=row, column=3, value=item["subject"][:50] if item["subject"] else "-").border = thin_border
            ws_all.cell(row=row, column=4, value=item["sender"][:30] if item["sender"] else "-").border = thin_border
            ws_all.cell(row=row, column=5, value=item["item_type"]).border = thin_border
            
            status_cell = ws_all.cell(row=row, column=6, value=item["status"].upper())
            status_cell.border = thin_border
            if item["status"] == "success":
                status_cell.fill = success_fill
            else:
                status_cell.fill = error_fill
            
            ws_all.cell(row=row, column=7, value=item["sedna_rec_id"] or "-").border = thin_border
            ws_all.cell(row=row, column=8, value=item["error_message"][:50] if item["error_message"] else "-").border = thin_border
            ws_all.cell(row=row, column=9, value=item["processed_at"].strftime("%Y-%m-%d %H:%M") if item["processed_at"] else "-").border = thin_border
        
        # Auto-fit column widths
        column_widths = [5, 10, 40, 25, 12, 10, 12, 40, 18]
        for i, width in enumerate(column_widths, start=1):
            ws_all.column_dimensions[get_column_letter(i)].width = width
        
        # =======================================================================
        # Successful Sheet
        # =======================================================================
        ws_success = wb.create_sheet("Successful")
        
        success_headers = ["#", "Email ID", "Subject", "Voucher No", "Type", "Sedna ID", "Processed At"]
        for col, header in enumerate(success_headers, start=1):
            cell = ws_success.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        success_items = [i for i in items if i["status"] == "success"]
        for row, item in enumerate(success_items, start=2):
            ws_success.cell(row=row, column=1, value=row - 1).border = thin_border
            ws_success.cell(row=row, column=2, value=item["email_id"]).border = thin_border
            ws_success.cell(row=row, column=3, value=item["subject"][:50] if item["subject"] else "-").border = thin_border
            ws_success.cell(row=row, column=4, value=item["voucher_no"] or "-").border = thin_border
            ws_success.cell(row=row, column=5, value=item["item_type"]).border = thin_border
            ws_success.cell(row=row, column=6, value=item["sedna_rec_id"] or "-").border = thin_border
            ws_success.cell(row=row, column=7, value=item["processed_at"].strftime("%Y-%m-%d %H:%M") if item["processed_at"] else "-").border = thin_border
        
        success_widths = [5, 10, 40, 15, 12, 12, 18]
        for i, width in enumerate(success_widths, start=1):
            ws_success.column_dimensions[get_column_letter(i)].width = width
        
        # =======================================================================
        # Failed Sheet
        # =======================================================================
        ws_failed = wb.create_sheet("Failed")
        
        failed_headers = ["#", "Email ID", "Subject", "Sender", "Type", "Error Message"]
        for col, header in enumerate(failed_headers, start=1):
            cell = ws_failed.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        failed_items = [i for i in items if i["status"] == "failed"]
        for row, item in enumerate(failed_items, start=2):
            ws_failed.cell(row=row, column=1, value=row - 1).border = thin_border
            ws_failed.cell(row=row, column=2, value=item["email_id"]).border = thin_border
            ws_failed.cell(row=row, column=3, value=item["subject"][:50] if item["subject"] else "-").border = thin_border
            ws_failed.cell(row=row, column=4, value=item["sender"][:30] if item["sender"] else "-").border = thin_border
            ws_failed.cell(row=row, column=5, value=item["item_type"]).border = thin_border
            ws_failed.cell(row=row, column=6, value=item["error_message"] or "-").border = thin_border
        
        failed_widths = [5, 10, 40, 25, 12, 60]
        for i, width in enumerate(failed_widths, start=1):
            ws_failed.column_dimensions[get_column_letter(i)].width = width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()


# Module-level service instance
_report_service: Optional[SyncReportService] = None


def set_report_service(service: SyncReportService):
    """Set the report service instance."""
    global _report_service
    _report_service = service


def get_report_service() -> SyncReportService:
    """Get the report service instance."""
    if not _report_service:
        raise RuntimeError("Report service not initialized")
    return _report_service
